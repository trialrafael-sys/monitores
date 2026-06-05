#!/usr/bin/env python3
"""
parse.py - Serasa Monitore TXT → data.json compacto
Uso: python parse.py arquivo1.txt [arquivo2.txt ...]
     python parse.py pasta/
     python parse.py *.TXT --export

Arquivos opcionais na mesma pasta:
  apelidos.xlsx  → CNPJ | Apelido | Observações
  grupos.xlsx    → CNPJ | Grupo | Segmento | Observações
"""

import sys, os, json, glob, re
from datetime import date as dtdate

MONTHS = {'JAN':'01','FEV':'02','MAR':'03','ABR':'04','MAI':'05','JUN':'06',
          'JUL':'07','AGO':'08','SET':'09','OUT':'10','NOV':'11','DEZ':'12'}

def clean_cnpj(s):
    """Remove pontuação do CNPJ e retorna só dígitos."""
    return re.sub(r'\D', '', str(s or ''))

def load_excel_map(path, key_col=0, val_col=1):
    """Lê Excel e retorna dict {cnpj_digits: valor}. Tolerante a erros."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        result = {}
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0: continue  # header
            if not row[key_col]: continue
            cnpj = clean_cnpj(row[key_col])
            if len(cnpj) >= 8:
                val = str(row[val_col] or '').strip()
                if val:
                    result[cnpj[:8]] = val  # use 8-digit root CNPJ as key
        return result
    except Exception as e:
        print(f"  Aviso: não foi possível ler {path}: {e}")
        return {}

def load_cedentes(path):
    """Lê cedentes.xlsx (coluna CNPJ, Apelido Principal, Grupo).
    Retorna (apelidos_dict, grupos_dict) ambos {cnpj8: valor}."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        apelidos, grupos = {}, {}
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0: continue
            if not row[0]: continue
            cnpj = clean_cnpj(row[0])
            if len(cnpj) < 8: continue
            key = cnpj[:8]
            apelido = str(row[1] or '').strip()
            grupo   = str(row[3] or '').strip()  # col D = Grupo
            if apelido: apelidos[key] = apelido
            if grupo and grupo.upper() not in ('NAN','NONE',''): grupos[key] = grupo
        return apelidos, grupos
    except Exception as e:
        print(f"  Aviso: não foi possível ler {path}: {e}")
        return {}, {}

def load_grupos(path):
    _, g = load_cedentes(path)
    return g

def load_apelidos(path):
    a, _ = load_cedentes(path)
    return a

def pmdate(s):
    s = s.strip()
    if len(s) >= 7 and s[:3].upper() in MONTHS:
        return MONTHS[s[:3].upper()] + '/20' + s[5:7]
    return s

def classify(t):
    t = t.upper()
    if 'PROTESTO' in t:      return 'PROTESTO'
    if 'DIVIDA VENCIDA' in t: return 'DIVIDA_VENCIDA'
    if 'ACAO JUDICIAL' in t:  return 'ACAO_JUDICIAL'
    if 'FALEN' in t:          return 'FALENCIA'
    if 'RECUP' in t:          return 'RECUPERACAO'
    if 'CHEQUE' in t:         return 'CHEQUE'
    return 'OUTROS'

FINANCIADOR_KEYWORDS = [
    'FUNDO', 'INVESTIMENTOS', 'FACTORING', 'SECURITIZADORA',
    'BANCO', 'CREDITORIOS', 'FINANCEIRA', 'FIDC',
    'CREDITO', 'CAPITAL', 'ASSET'
]

def is_financiador(credor):
    c = credor.upper()
    return any(k in c for k in FINANCIADOR_KEYWORDS)

def clean_credor(raw):
    """Remove prefixo numérico/documental antes do nome do credor."""
    import re as _re
    raw = raw.strip()
    if not raw: return raw

    def looks_like_name(s):
        return bool(_re.search(r'[A-Z]{3,}', s.strip())) and len(s.strip()) >= 4

    m = _re.match(r'^[\d./\-]+\s{2,}(.+)', raw)
    if m and looks_like_name(m.group(1)): return m.group(1).strip()
    m = _re.match(r'^[\d./\-]{8,}\s(.+)', raw)
    if m and looks_like_name(m.group(1)): return m.group(1).strip()
    m = _re.match(r'^[\d./\-]{6,}([A-Z].+)', raw)
    if m and looks_like_name(m.group(1)): return m.group(1).strip()
    m = _re.match(r'^(\d{{4,7}})\s([A-Z].+)', raw)
    if m and looks_like_name(m.group(2)): return m.group(2).strip()
    m = _re.match(r'^([A-Z0-9./\-]+)\s{{2,}}([A-Z].+)', raw)
    if m and _re.search(r'\d', m.group(1)) and looks_like_name(m.group(2)):
        return m.group(2).strip()
    m = _re.match(r'^([A-Z0-9]{{4,}})\s([A-Z].+)', raw)
    if m and _re.search(r'\d', m.group(1)) and looks_like_name(m.group(2)):
        return m.group(2).strip()
    return raw


def fname_to_date(path):
    name = os.path.basename(path)
    m = re.search(r'D(\d{2})(\d{2})(\d{2})', name)
    if m:
        return f'20{m.group(1)}-{m.group(2)}-{m.group(3)}'
    return dtdate.today().isoformat()

def parse_file(path):
    date = fname_to_date(path)
    pefin_val, refin_val = {}, {}

    with open(path, encoding='latin-1', errors='replace') as f:
        lines = f.readlines()

    # First pass: collect PEFIN/REFIN total values
    for line in lines:
        line = line.rstrip('\n\r')
        if len(line) < 35: continue
        cnpj = line[20:28]
        if not cnpj.isdigit(): continue
        seg = line[28:34]
        if seg in ('040101', '040102'):
            try: val = int(line[261:274].strip())
            except: val = 0
            if seg == '040101': pefin_val[cnpj] = val
            else: refin_val[cnpj] = val

    day_data = {}
    for line in lines:
        line = line.rstrip('\n\r')
        if len(line) < 35: continue
        cnpj = line[20:28]
        if not cnpj.isdigit(): continue
        seg = line[28:34]

        if cnpj not in day_data:
            day_data[cnpj] = {
                'razao': '', 'score': None,
                'pend': [], 'pefin_creds': set(), 'pefin': [],
                'refin_creds': set(), 'refin': [],
                'falrj': [], 'nc': False, 'nl': False, 'np': False, 'nr': False,
            }
        d = day_data[cnpj]

        if seg == '010102':
            r = line[34:114].strip()
            if r and not d['razao']: d['razao'] = r[:60]
        elif seg == '571001':
            try: d['score'] = int(line[34:49][6:9])
            except: pass
        elif seg == '041099':
            txt = line[28:]
            if 'NADA CONSTA PARA' in txt: d['nc'] = True
            elif 'NADA CONSTA' in txt:    d['nl'] = True
        elif seg == '040199' and 'NADA CONSTA' in line: d['np'] = True
        elif seg == '040299' and 'NADA CONSTA' in line: d['nr'] = True
        elif seg == '040202':
            tipo = line[43:70].strip()
            try: val = int(line[87:100].strip())
            except: val = 0
            d['pend'].append({
                'tipo': tipo, 'cat': classify(tipo),
                'ini': pmdate(line[70:77]), 'ult': pmdate(line[77:84]),
                'val': val, 'cred': line[100:122].strip(),
            })
        elif seg == '040101':
            cred = line[86:116].strip()
            key = cred[:20]
            if key not in d['pefin_creds']:
                d['pefin_creds'].add(key)
                d['pefin'].append({
                    't': line[60:73].strip() or 'PEFIN',
                    'c': cred, 'd': line[52:60].strip(),
                    'v': pefin_val.get(cnpj, 0),
                    'fin': is_financiador(cred),
                })
        elif seg == '040102':
            cr_raw = line[86:].rstrip()
            cr = clean_credor(cr_raw[:60])
            key = cr_raw[:20]
            try: val_item = int(line[73:86].strip())
            except: val_item = 0
            dt_item = f"{line[58:60]}/{line[56:58]}/{line[52:56]}" if line[52:60].strip().isdigit() else ''
            if key not in d['refin_creds']:
                d['refin_creds'].add(key)
                d['refin'].append({
                    't': line[60:73].strip() or 'REFIN',
                    'c': cr, 'dt': dt_item, 'v': val_item,
                    'fin': is_financiador(cr),
                })
        elif seg == '040601':
            d['falrj'].append({
                'tipo': line[51:72].strip(),
                'cidade': line[80:110].strip(),
                'data': line[43:51].strip(),
            })

    return date, day_data

def build_entry(snap):
    cats = {}
    for p in snap['pend']:
        cat = p['cat']
        if cat not in cats: cats[cat] = {'n': 0, 'v': 0, 'items': []}
        cats[cat]['n'] += 1
        cats[cat]['v'] += p['val']
        cats[cat]['items'].append({
            't': p['tipo'], 'i': p['ini'], 'u': p['ult'],
            'v': p['val'], 'c': p['cred'],
        })
    pv = sum(p.get('val_item', p.get('v', 0)) for p in snap['pefin'])
    rv = sum(p.get('val_item', p.get('v', 0)) for p in snap['refin'])
    any_fin_pefin = any(p.get('fin') for p in snap['pefin'])
    any_fin_refin = any(p.get('fin') for p in snap['refin'])
    return {
        's': snap['score'],
        'nc': snap['nc'], 'nl': snap['nl'],
        'np': snap['np'], 'nr': snap['nr'],
        'cats': cats,
        'pefin': snap['pefin'][:5],
        'refin': snap['refin'][:5],
        'falrj': snap['falrj'][:3],
        'tp': len(snap['pend']),
        'tpf': len(snap['pefin']), 'tpf_val': pv,
        'trf': len(snap['refin']), 'trf_val': rv,
        'tfr': len(snap['falrj']),
        'fin_pefin': any_fin_pefin,
        'fin_refin': any_fin_refin,
    }

def detect_changes(prev, curr, pd_, cd_):
    pc = set(prev['cats'].keys()); cc = set(curr['cats'].keys())
    new_cats = list(cc - pc); rem_cats = list(pc - cc)
    pfd = curr['tpf'] - prev['tpf']
    rfd = curr['trf'] - prev['trf']
    frd = curr['tfr'] - prev['tfr']
    sd = (curr['s'] - prev['s']) if (prev['s'] is not None and curr['s'] is not None and prev['s'] != curr['s']) else None
    lim = curr.get('nl', False)
    ma = []
    if pfd > 0 and prev['tpf'] == 0: ma.append('PEFIN_NOVO')
    elif pfd > 0: ma.append('PEFIN_AUMENTO')
    if rfd > 0 and prev['trf'] == 0: ma.append('REFIN_NOVO')
    elif rfd > 0: ma.append('REFIN_AUMENTO')
    if frd > 0: ma.append('FALRJ_NOVO')
    if lim: ma.append('LIMINAR')
    if curr.get('fin_pefin') and not prev.get('fin_pefin'): ma.append('PEFIN_FINANCIADOR')
    if not (new_cats or rem_cats or pfd or rfd or frd or sd or lim): return None
    ni, ri = [], []
    for cat in new_cats: ni.extend([{**i, 'cat': cat} for i in curr['cats'][cat]['items'][:3]])
    for cat in rem_cats: ri.extend([{**i, 'cat': cat} for i in prev['cats'][cat]['items'][:3]])
    if pfd > 0: ni.extend([{**p, 'cat': 'PEFIN'} for p in curr['pefin'][:2]])
    return {
        'f': pd_, 't': cd_,
        'nc': new_cats, 'rc': rem_cats,
        'pfd': pfd, 'rfd': rfd, 'frd': frd,
        'sd': sd, 'lim': lim, 'ma': ma,
        'ni': ni[:5], 'ri': ri[:5],
    }

def merge_into(all_data, cnpj_razao, date, day_data):
    for cnpj, snap in day_data.items():
        if cnpj not in all_data: all_data[cnpj] = {}
        if date in all_data[cnpj]:
            ex = all_data[cnpj][date]
            if snap['razao'] and not ex['razao']: ex['razao'] = snap['razao']
            if snap['score'] is not None: ex['score'] = snap['score']
            ex['pend'].extend(snap['pend'])
            if not ex['pefin']: ex['pefin'] = snap['pefin']
            if not ex['refin']: ex['refin'] = snap['refin']
            ex['falrj'].extend(snap['falrj'])
            if snap['nc']: ex['nc'] = True
            if snap['nl']: ex['nl'] = True
        else:
            all_data[cnpj][date] = snap
        if snap['razao']: cnpj_razao[cnpj] = snap['razao']

def build_output(all_data, cnpj_razao, apelidos, grupos):
    result = {}
    missing_apelido = []
    missing_grupo = []

    for cnpj, dates in all_data.items():
        razao = cnpj_razao.get(cnpj, cnpj)
        apelido = apelidos.get(cnpj, '')
        grupo = grupos.get(cnpj, '')

        if not apelido:
            missing_apelido.append(cnpj)
            apelido = razao[:40] + ' (falta atualização)'
        if not grupo:
            missing_grupo.append(cnpj)

        sd = sorted(dates.keys())
        tl, changes = {}, []
        for i, d in enumerate(sd):
            entry = build_entry(dates[d])
            tl[d] = entry
            if i > 0:
                ch = detect_changes(tl[sd[i-1]], entry, sd[i-1], d)
                if ch: changes.append(ch)

        result[cnpj] = {
            'r': razao[:50].strip(),
            'a': apelido[:50].strip(),
            'g': grupo[:50].strip(),
            'tl': tl,
            'ch': changes,
        }

    return result, missing_apelido, missing_grupo

def main():
    args = sys.argv[1:]
    export_mode = '--export' in args
    args = [a for a in args if a != '--export']

    if not args:
        print(__doc__)
        sys.exit(1)

    # Collect files
    files = []
    for a in args:
        if os.path.isdir(a):
            files += sorted(glob.glob(os.path.join(a, '*.TXT')) +
                           glob.glob(os.path.join(a, '*.txt')))
        elif '*' in a:
            files += sorted(glob.glob(a))
        else:
            files.append(a)
    files = list(dict.fromkeys(files))

    if not files:
        print("Nenhum arquivo .TXT encontrado.")
        sys.exit(1)

    # Load apelidos / grupos
    apelidos, grupos = {}, {}
    # Tenta cedentes.xlsx (arquivo único) primeiro, depois apelidos/grupos separados
    for fname in ['cedentes.xlsx', 'cedentes.xlsx'.upper()]:
        if os.path.exists(fname):
            apelidos, grupos = load_cedentes(fname)
            print(f"  Cedentes: {len(apelidos)} apelidos + {len(grupos)} grupos de {fname}")
            break
    else:
        for fname in ['apelidos.xlsx', 'apelidos.xlsx'.upper()]:
            if os.path.exists(fname):
                apelidos = load_apelidos(fname)
                print(f"  Apelidos: {len(apelidos)} carregados de {fname}")
                break
        for fname in ['grupos.xlsx', 'grupos.xlsx'.upper()]:
            if os.path.exists(fname):
                grupos = load_grupos(fname)
                print(f"  Grupos:   {len(grupos)} carregados de {fname}")
                break

    # Load existing data.json
    all_data, cnpj_razao, existing_dates = {}, {}, []
    if os.path.exists('data.json'):
        print("Carregando data.json existente...")
        with open('data.json', encoding='utf-8') as f:
            old = json.load(f)
        existing_dates = old.get('dates', [])
        for cnpj, cd in old.get('cnpjs', {}).items():
            cnpj_razao[cnpj] = cd.get('r', cnpj)
            all_data[cnpj] = {}
            for d, entry in cd.get('tl', {}).items():
                pend = []
                for cat, cdata in entry.get('cats', {}).items():
                    for item in cdata.get('items', []):
                        pend.append({'tipo': item.get('t',''), 'cat': cat,
                                     'ini': item.get('i',''), 'ult': item.get('u',''),
                                     'val': item.get('v',0), 'cred': item.get('c','')})
                all_data[cnpj][d] = {
                    'razao': cnpj_razao[cnpj], 'score': entry.get('s'),
                    'pend': pend, 'pefin_creds': set(),
                    'pefin': entry.get('pefin', []),
                    'refin_creds': set(),
                    'refin': entry.get('refin', []),
                    'falrj': entry.get('falrj', []),
                    'nc': entry.get('nc', False), 'nl': entry.get('nl', False),
                    'np': entry.get('np', False), 'nr': entry.get('nr', False),
                }
        print(f"  {len(all_data)} CNPJs, {len(existing_dates)} datas existentes")

    # Process new files
    new_dates = []
    for fpath in files:
        date, day_data = parse_file(fpath)
        if date not in existing_dates:
            print(f"  {os.path.basename(fpath)} → {date} ({len(day_data)} CNPJs)")
            merge_into(all_data, cnpj_razao, date, day_data)
            if date not in new_dates: new_dates.append(date)
        else:
            print(f"  Pulando {os.path.basename(fpath)} (data {date} já existe)")

    if not new_dates:
        print("Nenhuma data nova. data.json não alterado.")
        return

    result, miss_apelido, miss_grupo = build_output(all_data, cnpj_razao, apelidos, grupos)

    if miss_apelido:
        print(f"\n⚠  {len(miss_apelido)} CNPJs sem apelido em apelidos.xlsx:")
        for c in miss_apelido[:10]:
            print(f"     {c} — {cnpj_razao.get(c,'?')[:40]}")
        if len(miss_apelido) > 10:
            print(f"     ... e mais {len(miss_apelido)-10}")

    if miss_grupo:
        print(f"\n⚠  {len(miss_grupo)} CNPJs sem grupo em grupos.xlsx")

    all_dates = sorted(set(d for cd in result.values() for d in cd['tl']))
    output = {'dates': all_dates, 'cnpjs': result}

    with open('data.json', 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, separators=(',', ':'))

    sz = os.path.getsize('data.json')
    print(f"\ndata.json: {sz/1024:.0f} KB — {len(result)} CNPJs — {len(all_dates)} datas")

    if export_mode:
        hist = {'gerado': dtdate.today().isoformat(), 'cnpjs': {}}
        for cnpj, cd in output['cnpjs'].items():
            if cd['ch']:
                hist['cnpjs'][cnpj] = {'r': cd['r'], 'a': cd['a'], 'g': cd['g'], 'ch': cd['ch']}
        name = f"historico_{dtdate.today().strftime('%Y%m%d')}.json"
        with open(name, 'w', encoding='utf-8') as f:
            json.dump(hist, f, ensure_ascii=False, separators=(',', ':'))
        print(f"{name}: {os.path.getsize(name)/1024:.0f} KB (só mudanças)")

    print("Feito.")

if __name__ == '__main__':
    main()
